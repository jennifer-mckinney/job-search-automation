#!/usr/bin/env python3
"""
board_ingest.py - carries your cockpit card moves into the Applications Log.

The browser cannot silently write to disk, so one click has to carry the moves
across: you press "Export board_data.json" on the Board tab. This script does
everything after that click.

  1. Sweeps ~/Downloads for board_data*.json and moves it into _inbox/
     (Chrome downloads there by default; _inbox stays the canonical location).
  2. Reads the newest export.
  3. Updates the Applications Log - matching rows by COMPANY, never by row index.
  4. Rewrites the cockpit's SEED columns and stamps SEED.updatedAt, so your
     pending moves see the file agree with them and retire themselves.
  5. Archives the consumed export so it is never ingested twice.
  6. Logs every action to the run log.

Deliberately NOT a daemon. A long-running watcher dies silently and you never
find out. This is idempotent: safe to run at the top of every run, and from
launchd/cron every minute. Running it a thousand times does nothing a thousand
times.

    python3 board_ingest.py --once            # sweep, ingest, archive
    python3 board_ingest.py --once --dry-run  # show what it WOULD do

Copyright (C) 2026 Jennifer McKinney
SPDX-License-Identifier: GPL-3.0-or-later
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import runlog

HERE = Path(__file__).resolve().parent
INBOX = HERE / "_inbox"
ARCHIVE = INBOX / "archive"
DOWNLOADS = Path.home() / "Downloads"
COCKPIT = HERE.parent / "cockpit" / "Job_Search_Cockpit.html"
WORKBOOK = HERE.parent / "applications_log.xlsx"  # override with --workbook
SHEET = "Applications Log"

# Board column id -> Applications Log Status value.
COL_TO_STATUS: Dict[str, str] = {
    "sourced": "Researching",
    "applying": "Tailored - ready to apply",
    "applied": "Applied",
    "screen": "Screen",
    "interview": "Interview",
    "offer": "Offer",
    "closed": "Closed",
}


def utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def sweep_downloads(downloads: Path, inbox: Path, dry: bool = False) -> List[Path]:
    """Move any board_data*.json out of Downloads and into the inbox."""
    moved: List[Path] = []
    if not downloads.exists():
        return moved
    for f in sorted(downloads.glob("board_data*.json")):
        dest = inbox / f"board_data_{int(f.stat().st_mtime)}.json"
        if not dry:
            inbox.mkdir(parents=True, exist_ok=True)
            shutil.move(str(f), str(dest))
        moved.append(dest)
    return moved


def newest_export(inbox: Path) -> Optional[Path]:
    files = [f for f in inbox.glob("board_data*.json") if f.is_file()]
    return max(files, key=lambda f: f.stat().st_mtime) if files else None


def load_export(path: Path) -> List[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    cards = data.get("cards")
    if not isinstance(cards, list):
        raise ValueError(f"{path.name}: no 'cards' array - not a board export")
    for c in cards:
        if "company" not in c or "column" not in c:
            raise ValueError(f"{path.name}: a card is missing company/column")
    return cards


def update_workbook(wb_path: Path, cards: List[dict], sheet: str = SHEET,
                    dry: bool = False) -> List[Tuple[str, str, str]]:
    """Update Status by matching COMPANY in column B. Never by row index.

    Returns [(company, old_status, new_status), ...] for rows that changed.
    """
    import openpyxl
    wb = openpyxl.load_workbook(wb_path)
    ws = wb[sheet]

    # Find the header row and the Company / Status columns by NAME.
    hdr_row = col_company = col_status = col_notes = None
    for r in range(1, min(ws.max_row, 20) + 1):
        vals = {str(ws.cell(r, c).value).strip(): c
                for c in range(1, ws.max_column + 1) if ws.cell(r, c).value}
        if "Company" in vals and "Status" in vals:
            hdr_row, col_company, col_status = r, vals["Company"], vals["Status"]
            col_notes = vals.get("Notes")
            break
    if not hdr_row:
        raise ValueError(f"{wb_path.name}: no header row with Company + Status")

    by_company = {str(c["company"]).strip().lower(): c for c in cards}
    changed: List[Tuple[str, str, str]] = []

    for r in range(hdr_row + 1, ws.max_row + 1):
        cell_co = ws.cell(r, col_company).value
        if not cell_co:
            continue
        card = by_company.get(str(cell_co).strip().lower())
        if not card:
            continue
        new = COL_TO_STATUS.get(card["column"])
        if not new:
            continue
        if card.get("flag") and "on hold" in str(card["flag"]).lower():
            new = "On hold (Workday)"
        old = str(ws.cell(r, col_status).value or "").strip()
        if old == new:
            continue
        changed.append((str(cell_co), old, new))
        if not dry:
            ws.cell(r, col_status).value = new
            if col_notes:
                note = str(ws.cell(r, col_notes).value or "")
                stamp = (f"Status -> {new} on {utc()[:10]} "
                         f"(ingested from cockpit board export).")
                ws.cell(r, col_notes).value = (note + " " + stamp).strip()

    if changed and not dry:
        wb.save(wb_path)
    return changed


def update_cockpit(cockpit: Path, cards: List[dict], dry: bool = False) -> List[str]:
    """Rewrite SEED card columns + stamp SEED.updatedAt.

    This is what makes her pending moves retire: the file finally agrees with them.
    """
    html = cockpit.read_text(encoding="utf-8")
    updated: List[str] = []

    for c in cards:
        company, col = str(c["company"]), str(c["column"])
        # Find this card's object in SEED by company, then set its col.
        pat = re.compile(
            r'(\{id:"[^"]+",company:"' + re.escape(company) + r'".*?col:")([a-z]+)(")',
            re.S)
        m = pat.search(html)
        if not m or m.group(2) == col:
            continue
        html = html[:m.start()] + m.group(1) + col + m.group(3) + html[m.end():]
        updated.append(f"{company}: {m.group(2)} -> {col}")

    if updated:
        html = re.sub(r'updatedAt:"[^"]*"', f'updatedAt:"{utc()}"', html, count=1)
        html = re.sub(r'(const SEED=\{updated:")[^"]*(")',
                      r"\g<1>" + utc()[:10] + r"\2", html, count=1)
        if not dry:
            cockpit.write_text(html, encoding="utf-8")
    return updated


def ingest(run_id: str, inbox: Path = INBOX, downloads: Path = DOWNLOADS,
           cockpit: Path = COCKPIT, workbook: Path = WORKBOOK,
           log_dir: Optional[Path] = None, dry: bool = False) -> int:
    log_dir = log_dir or runlog.default_log_dir()

    def log(event, status, detail, meta=None):
        try:
            runlog.append_event(log_dir, run_id, "P5b", event, status,
                                detail=detail, meta=meta or {})
        except Exception as exc:
            print(f"warning: run log write failed: {exc}", file=sys.stderr)

    inbox.mkdir(parents=True, exist_ok=True)
    swept = sweep_downloads(downloads, inbox, dry)
    if swept:
        log("inbox_sweep", "ok", f"moved {len(swept)} export(s) out of Downloads",
            {"count": len(swept)})

    export = newest_export(inbox)
    if not export:
        log("board_ingest", "skip", "no board export waiting - nothing to reconcile")
        print("nothing to ingest")
        return 0

    try:
        cards = load_export(export)
    except ValueError as exc:
        log("board_ingest", "error", str(exc))
        print(f"error: {exc}", file=sys.stderr)
        return 1

    changed = update_workbook(workbook, cards, dry=dry)
    seeded = update_cockpit(cockpit, cards, dry=dry)

    for co, old, new in changed:
        log("log_status_update", "ok", f"{co}: {old or '(blank)'} -> {new}",
            {"company": co, "from": old, "to": new})
    if seeded:
        log("cockpit_seed_sync", "ok", "; ".join(seeded), {"cards": len(seeded)})

    if not dry:
        archive = inbox / "archive"          # relative to THIS inbox, not the global
        archive.mkdir(parents=True, exist_ok=True)
        shutil.move(str(export),
                    str(archive / f"{export.stem}_{utc()[:19].replace(':', '')}.json"))

    log("board_ingest", "ok" if (changed or seeded) else "skip",
        f"{len(changed)} log row(s), {len(seeded)} cockpit card(s) updated",
        {"log_rows": len(changed), "cockpit_cards": len(seeded)})

    print(f"{'DRY RUN: ' if dry else ''}ingested {export.name}")
    for co, old, new in changed:
        print(f"  log:     {co}: {old or '(blank)'} -> {new}")
    for s in seeded:
        print(f"  cockpit: {s}")
    if not changed and not seeded:
        print("  (already in sync - nothing to do)")
    return 0


def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        prog="board_ingest.py",
        description="Reconcile an exported cockpit board into the Applications Log.")
    p.add_argument("--once", action="store_true", default=True,
                   help="sweep, ingest, archive (default)")
    p.add_argument("--dry-run", action="store_true", help="show what would change")
    p.add_argument("--run-id", help="attach to an existing run (default: new id)")
    p.add_argument("--inbox", type=Path, default=INBOX)
    p.add_argument("--downloads", type=Path, default=DOWNLOADS)
    p.add_argument("--cockpit", type=Path, default=COCKPIT)
    p.add_argument("--workbook", type=Path, default=WORKBOOK)
    a = p.parse_args(argv)
    run_id = a.run_id or runlog.new_run_id()
    return ingest(run_id, a.inbox, a.downloads, a.cockpit, a.workbook, dry=a.dry_run)


if __name__ == "__main__":
    raise SystemExit(main())
