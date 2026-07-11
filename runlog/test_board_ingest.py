#!/usr/bin/env python3
"""
Full-loop test: drag in a REAL browser -> Export -> ingest -> log + cockpit updated
-> the pending move retires itself on the next load.

This is the loop that closes R7 (the cockpit running ahead of the log).

    python3 test_board_ingest.py

Copyright (C) 2026 Jennifer McKinney
SPDX-License-Identifier: GPL-3.0-or-later
"""
from __future__ import annotations

import functools
import http.server
import shutil
import socketserver
import tempfile
import threading
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

import board_ingest as BI
import runlog

PASS = FAIL = 0


def chk(name, cond, extra=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  PASS  {name}")
    else:
        FAIL += 1
        print(f"  FAIL  {name}  {extra}")


def make_workbook(path: Path, companies):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = BI.SHEET
    ws["A1"] = "Applications Log"
    ws.append([])
    ws.append(["Date Applied", "Company", "Role Title", "Target Role Type",
               "Resume Variant", "Skills", "Projects", "Job URL", "Status", "Notes"])
    for co in companies:
        ws.append(["2026-07-10", co, "Role", "R4", "V2", "s", "p", "u", "Applied", ""])
    wb.save(path)


def status_of(path: Path, company: str) -> str:
    ws = openpyxl.load_workbook(path)[BI.SHEET]
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value).strip().lower() == company.lower():
            return str(ws.cell(r, 9).value or "")
    return ""


def main() -> int:
    tmp = Path(tempfile.mkdtemp())
    cockpit = tmp / "cockpit.html"
    shutil.copy("Job_Search_Cockpit.html", cockpit)
    inbox, downloads, logs = tmp / "_inbox", tmp / "Downloads", tmp / "logs"
    for d in (inbox, downloads, logs):
        d.mkdir(parents=True)
    wb = tmp / "log.xlsx"

    # Seed the workbook from the cockpit's own SEED companies.
    import re
    companies = re.findall(r'company:"([^"]+)"', cockpit.read_text())
    make_workbook(wb, companies)
    target = companies[0]
    chk("fixture: workbook starts with the card as Applied",
        status_of(wb, target) == "Applied", status_of(wb, target))

    # --- serve + drive a REAL browser ------------------------------------
    handler = functools.partial(http.server.SimpleHTTPRequestHandler,
                                directory=str(tmp))
    httpd = socketserver.TCPServer(("127.0.0.1", 0), handler)
    threading.Thread(target=httpd.serve_forever, daemon=True).start()
    url = f"http://127.0.0.1:{httpd.server_address[1]}/cockpit.html"

    with sync_playwright() as p:
        b = p.chromium.launch(channel="chrome", headless=True)
        ctx = b.new_context(accept_downloads=True)
        pg = ctx.new_page()
        pg.goto(url)
        pg.wait_for_selector("#c_board .col", state="attached")

        cid = pg.evaluate(f"SEED.cards.find(c => c.company === {target!r}).id")
        pg.evaluate("""([id]) => {
            const co = STATE.cards.find(c => c.id === id).company;
            const src = [...document.querySelectorAll('#c_board .kcard')]
                .find(c => c.querySelector('.kco').textContent.trim() === co);
            const dst = [...document.querySelectorAll('#c_board .col')]
                .find(c => c.querySelector('.colhead .n').textContent.trim() === 'Screening');
            const dt = new DataTransfer();
            src.dispatchEvent(new DragEvent('dragstart', {bubbles:true, dataTransfer:dt}));
            dst.dispatchEvent(new DragEvent('drop', {bubbles:true, dataTransfer:dt}));
        }""", [cid])
        pg.wait_for_timeout(150)
        chk("real drag: card moved to Screening in the browser",
            pg.evaluate(f"STATE.cards.find(c=>c.id==={cid!r}).col") == "screen")
        chk("real drag: move is pending (not yet in the log)",
            pg.evaluate("JSON.parse(localStorage.getItem('cockpit_pending_v2')||'[]').length") == 1)

        # She clicks Export -> lands in "Downloads"
        pg.click("nav button[data-t='board']")
        with pg.expect_download() as dl:
            pg.click("#c_export")
        dl.value.save_as(str(downloads / "board_data.json"))
        chk("Export produced a file in Downloads",
            (downloads / "board_data.json").exists())

        # --- THE INGEST -------------------------------------------------
        rid = runlog.new_run_id()
        rc = BI.ingest(rid, inbox=inbox, downloads=downloads, cockpit=cockpit,
                       workbook=wb, log_dir=logs)
        chk("ingest exits clean", rc == 0)
        chk("Downloads was swept (file no longer there)",
            not (downloads / "board_data.json").exists())
        chk("export was archived, so it cannot be ingested twice",
            any((inbox / "archive").glob("*.json")))
        chk("APPLICATIONS LOG updated: Applied -> Screen",
            status_of(wb, target) == "Screen", status_of(wb, target))
        chk("COCKPIT SEED updated to screen",
            f'col:"screen"' in cockpit.read_text())

        # --- the whole point: the pending move now retires itself --------
        pg.reload()
        pg.wait_for_selector("#c_board .col", state="attached")
        chk("card still shows Screening after the run logged it",
            pg.evaluate(f"STATE.cards.find(c=>c.id==={cid!r}).col") == "screen")
        chk("PENDING MOVE RETIRED ITSELF (log and cockpit now agree)",
            pg.evaluate("JSON.parse(localStorage.getItem('cockpit_pending_v2')||'[]').length") == 0)
        b.close()

    # --- idempotency ----------------------------------------------------
    rc = BI.ingest(runlog.new_run_id(), inbox=inbox, downloads=downloads,
                   cockpit=cockpit, workbook=wb, log_dir=logs)
    chk("running ingest again is a safe no-op", rc == 0)
    chk("status unchanged after a second ingest", status_of(wb, target) == "Screen")

    evs = runlog.read_events(logs)
    chk("status change is in the run log",
        any(e["event"] == "log_status_update" for e in evs))
    chk("a no-op ingest still logs a skip",
        any(e["event"] == "board_ingest" and e["status"] == "skip" for e in evs))

    httpd.shutdown()
    shutil.rmtree(tmp, ignore_errors=True)
    print(f"\n{PASS} passed, {FAIL} failed")
    return 1 if FAIL else 0


if __name__ == "__main__":
    raise SystemExit(main())
